#!/usr/bin/env python3
"""
Скрипт для извлечения данных из PDF спецификаций и создания Excel файла
"""
import os
import re
import pandas as pd
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import math

# Константы для расчета массы труб по ГОСТ 18599-2001
# Плотность полиэтилена ПЭ100 = 0.96 г/см³
PE_DENSITY = 0.96  # г/см³


def calculate_pipe_mass_per_meter(diameter, wall_thickness):
    """
    Расчет массы трубы за 1 метр по ГОСТ 18599-2001
    
    Формула: m = π * (D - e) * e * ρ / 1000
    где:
    m - масса 1 метра трубы (кг/м)
    D - наружный диаметр (мм)
    e - толщина стенки (мм)
    ρ - плотность материала (г/см³)
    π - 3.14159
    """
    # Формула массы трубы
    mass = math.pi * (diameter - wall_thickness) * wall_thickness * PE_DENSITY / 1000
    return round(mass, 2)


def extract_pipe_parameters(nomenclature):
    """
    Извлечение параметров трубы из номенклатуры
    Формат: "Труба ... SDR17 - 160 х9,50" или "Труба ... ∅160х23,7"
    
    Возвращает: (диаметр, толщина стенки) или None
    """
    # Паттерн 1: SDR формат (160 х9,50)
    pattern1 = r'(\d+)\s*[хx×]\s*(\d+[,.]?\d*)'
    # Паттерн 2: ∅ формат (∅160х23,7)
    pattern2 = r'[∅Ø](\d+)[хx×](\d+[,.]?\d*)'
    
    match = re.search(pattern1, nomenclature) or re.search(pattern2, nomenclature)
    
    if match:
        diameter = int(match.group(1))
        thickness = float(match.group(2).replace(',', '.'))
        return diameter, thickness
    
    return None


def is_pipe(nomenclature):
    """Проверка, является ли номенклатура трубой"""
    if not nomenclature or pd.isna(nomenclature):
        return False
    nomenclature_lower = str(nomenclature).lower()
    return 'труба' in nomenclature_lower or 'футляр' in nomenclature_lower


def is_pipe_or_fitting(nomenclature):
    """Проверка, является ли номенклатура трубой или фитингом"""
    if not nomenclature or pd.isna(nomenclature):
        return False
    
    nomenclature_lower = str(nomenclature).lower()
    
    # Трубы
    if 'труба' in nomenclature_lower or 'футляр' in nomenclature_lower:
        return True
    
    # Фитинги (только указанные)
    fittings = ['муфта', 'отвод', 'втулка', 'фланец']
    
    return any(fitting in nomenclature_lower for fitting in fittings)


def parse_pdf_file(pdf_path):
    """
    Парсинг одного PDF файла и извлечение данных
    
    Возвращает список словарей с данными:
    [{
        'Файл': 'имя файла',
        'Номенклатура': '...',
        'Количество': ...,
        'Масса': ...,
        'Завод изготовитель': '...'
    }, ...]
    """
    filename = os.path.basename(pdf_path)
    # Убираем расширение .pdf из имени файла
    filename_without_ext = filename.replace('.pdf', '')
    data = []
    
    print(f"\n📄 Обрабатываю: {filename}")
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages, 1):
                # Извлекаем таблицы со страницы
                tables = page.extract_tables()
                
                if not tables:
                    continue
                
                for table in tables:
                    if not table or len(table) < 2:
                        continue
                    
                    # Ищем заголовки столбцов
                    header_row = None
                    header_idx = -1
                    
                    for idx, row in enumerate(table[:5]):  # Проверяем первые 5 строк
                        if row and any('Наименование' in str(cell) for cell in row if cell):
                            header_row = row
                            header_idx = idx
                            break
                    
                    if not header_row:
                        continue
                    
                    # Определяем индексы нужных столбцов
                    name_col = None
                    qty_col = None
                    manufacturer_col = None
                    
                    for col_idx, header in enumerate(header_row):
                        if not header:
                            continue
                        header_str = str(header).strip().replace('\n', '').replace(' ', '').replace('-', '')
                        
                        if 'Наименование' in str(header) and 'техническ' in str(header).lower():
                            name_col = col_idx
                        elif 'колич' in header_str.lower() or 'количество' in header_str.lower():
                            qty_col = col_idx
                        elif 'Завод' in str(header) or 'изготовитель' in str(header).lower():
                            manufacturer_col = col_idx
                    
                    # Извлекаем данные из строк таблицы
                    for row in table[header_idx + 1:]:
                        if not row or len(row) <= max(filter(None, [name_col, qty_col, manufacturer_col])):
                            continue
                        
                        nomenclature = row[name_col] if name_col is not None else ''
                        quantity = row[qty_col] if qty_col is not None else ''
                        manufacturer = row[manufacturer_col] if manufacturer_col is not None else ''
                        
                        # Пропускаем пустые строки
                        if not nomenclature or str(nomenclature).strip() == '':
                            continue
                        
                        # Очистка данных
                        nomenclature = str(nomenclature).strip()
                        
                        # Фильтр: оставляем только трубы и фитинги
                        if not is_pipe_or_fitting(nomenclature):
                            continue
                        
                        # Парсинг количества
                        try:
                            if quantity and str(quantity).strip():
                                quantity_str = str(quantity).strip().replace(',', '.').replace(' ', '')
                                # Извлекаем число из строки
                                qty_match = re.search(r'(\d+[.,]?\d*)', quantity_str)
                                if qty_match:
                                    quantity = float(qty_match.group(1).replace(',', '.'))
                                else:
                                    quantity = None
                            else:
                                quantity = None
                        except:
                            quantity = None
                        
                        # Очистка производителя
                        manufacturer = str(manufacturer).strip() if manufacturer else ''
                        
                        # Расчет массы для труб
                        mass = None
                        if is_pipe(nomenclature):
                            params = extract_pipe_parameters(nomenclature)
                            if params:
                                diameter, thickness = params
                                mass = calculate_pipe_mass_per_meter(diameter, thickness)
                        
                        data.append({
                            'Файл': filename_without_ext,
                            'Номенклатура': nomenclature,
                            'Количество': quantity,
                            'Масса': mass,
                            'Завод изготовитель': manufacturer
                        })
    
    except Exception as e:
        print(f"   ⚠️  Ошибка при обработке {filename}: {e}")
    
    print(f"   ✓ Извлечено строк: {len(data)}")
    return data


def create_excel_with_formatting(data, output_file):
    """
    Создание Excel файла с форматированием
    - Серые строки для разделения файлов
    - Выравнивание и границы
    """
    if not data:
        print("⚠️  Нет данных для записи")
        return
    
    # Создаем DataFrame
    df = pd.DataFrame(data)
    
    # Создаем Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Спецификация"
    
    # Стили
    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Заголовки
    headers = ['Файл', 'Номенклатура', 'Количество', 'Масса', 'Завод изготовитель']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
    
    # Данные
    current_file = None
    row_idx = 2
    
    for idx, row in df.iterrows():
        # Если новый файл - добавляем строку-разделитель
        if row['Файл'] != current_file:
            current_file = row['Файл']
            
            # Строка с названием файла (серый фон)
            if row_idx > 2:  # Не добавляем пустую строку перед первым файлом
                row_idx += 1
            
            file_cell = ws.cell(row=row_idx, column=1, value=current_file)
            file_cell.fill = gray_fill
            file_cell.font = bold_font
            file_cell.alignment = left_align
            file_cell.border = thin_border
            
            # Объединяем ячейки для названия файла
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=5)
            
            # Применяем стили ко всем ячейкам объединенной строки
            for col in range(1, 6):
                cell = ws.cell(row=row_idx, column=col)
                cell.fill = gray_fill
                cell.border = thin_border
            
            row_idx += 1
        
        # Данные строки
        ws.cell(row=row_idx, column=1, value='').border = thin_border
        ws.cell(row=row_idx, column=2, value=row['Номенклатура']).border = thin_border
        ws.cell(row=row_idx, column=2).alignment = left_align
        
        qty_cell = ws.cell(row=row_idx, column=3, value=row['Количество'] if pd.notna(row['Количество']) else '')
        qty_cell.border = thin_border
        qty_cell.alignment = center_align
        
        mass_cell = ws.cell(row=row_idx, column=4, value=row['Масса'] if pd.notna(row['Масса']) else '')
        mass_cell.border = thin_border
        mass_cell.alignment = center_align
        
        manuf_cell = ws.cell(row=row_idx, column=5, value=row['Завод изготовитель'])
        manuf_cell.border = thin_border
        manuf_cell.alignment = center_align
        
        row_idx += 1
    
    # Настройка ширины столбцов
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 70
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 25
    
    # Закрепление первой строки
    ws.freeze_panes = 'A2'
    
    # Сохранение
    wb.save(output_file)
    print(f"\n✅ Excel файл сохранен: {output_file}")
    print(f"   Всего строк: {len(df)}")


def main():
    """Основная функция"""
    print("=" * 80)
    print("🚀 ПАРСИНГ PDF СПЕЦИФИКАЦИЙ")
    print("=" * 80)
    
    # Путь к папке с PDF
    pdf_folder = '/Users/exmuzzy2/git/fedor/pdf'
    
    # Получаем список всех PDF файлов
    pdf_files = [f for f in os.listdir(pdf_folder) if f.endswith('.pdf')]
    pdf_files.sort()  # Сортируем по имени
    
    print(f"\n📁 Найдено PDF файлов: {len(pdf_files)}")
    
    # Парсим все файлы
    all_data = []
    
    for pdf_file in pdf_files:
        pdf_path = os.path.join(pdf_folder, pdf_file)
        file_data = parse_pdf_file(pdf_path)
        all_data.extend(file_data)
    
    print(f"\n📊 Всего извлечено записей: {len(all_data)}")
    
    # Создаем Excel
    output_file = '/Users/exmuzzy2/git/fedor/specifications_full.xlsx'
    create_excel_with_formatting(all_data, output_file)
    
    print("\n" + "=" * 80)
    print("✅ ГОТОВО!")
    print("=" * 80)


if __name__ == '__main__':
    main()

